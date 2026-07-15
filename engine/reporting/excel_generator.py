"""Excel report generator (M14).

Generates Excel reports from report payload using openpyxl.
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

from config.constants import AuditCategory, AuditSeverity
from models.domain import AuditEntry, AuditLog, ReportPackage


def generate_excel_report(
    report_package: ReportPackage,
    output_path: str,
    audit_log: AuditLog | None = None,
) -> str:
    """Generate Excel report from report payload.

    Args:
        report_package: Complete report with all 13 sections
        output_path: Path to write Excel file
        audit_log: Optional audit log

    Returns:
        Path to generated Excel file
    """
    if audit_log is None:
        audit_log = AuditLog()

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        audit_log.add(AuditEntry(
            timestamp=datetime.now(),
            module_name="excel_generator",
            action="generate_excel_report",
            decision="WARNING",
            reason="openpyxl not installed — writing CSV fallback instead",
            severity=AuditSeverity.WARNING,
            category=AuditCategory.PIPELINE,
        ))
        # Fallback: write CSV
        csv_path = str(Path(output_path).with_suffix(".csv"))
        _write_csv_fallback(report_package, csv_path)
        return csv_path

    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Create a sheet for each section
    for section_name, section_data in report_package.sections.items():
        title = section_data.get("title", section_name)[:31]  # Excel sheet name limit
        ws = wb.create_sheet(title=title)
        
        # Header style
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Write section title
        ws["A1"] = title
        ws["A1"].font = Font(bold=True, size=14)
        
        row = 3
        for key, value in section_data.items():
            if key == "title":
                continue
                
            if isinstance(value, dict):
                ws.cell(row=row, column=1, value=key).font = header_font
                row += 1
                for sub_key, sub_value in value.items():
                    ws.cell(row=row, column=1, value=f"  {sub_key}")
                    ws.cell(row=row, column=2, value=str(sub_value))
                    row += 1
                row += 1
                
            elif isinstance(value, list):
                ws.cell(row=row, column=1, value=key).font = header_font
                row += 1
                
                if value and isinstance(value[0], dict):
                    # Table header
                    headers = list(value[0].keys())
                    for col_idx, header in enumerate(headers, start=1):
                        cell = ws.cell(row=row, column=col_idx, value=header)
                        cell.font = header_font
                        cell.fill = header_fill
                    row += 1
                    
                    # Table rows
                    for item in value:
                        for col_idx, header in enumerate(headers, start=1):
                            ws.cell(row=row, column=col_idx, value=str(item.get(header, "")))
                        row += 1
                else:
                    for item in value:
                        ws.cell(row=row, column=1, value=str(item))
                        row += 1
                row += 1
                
            else:
                ws.cell(row=row, column=1, value=key).font = header_font
                ws.cell(row=row, column=2, value=str(value))
                row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(output_path)

    audit_log.add(AuditEntry(
        timestamp=datetime.now(),
        module_name="excel_generator",
        action="generate_excel_report",
        decision="SUCCESS",
        reason=f"Excel report written to {output_path}",
        severity=AuditSeverity.INFO,
        category=AuditCategory.PIPELINE,
    ))

    return output_path


def _write_csv_fallback(report_package: ReportPackage, csv_path: str) -> None:
    """Write CSV fallback if openpyxl not available."""
    import csv
    
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        
        for section_name, section_data in report_package.sections.items():
            title = section_data.get("title", section_name)
            writer.writerow([title])
            writer.writerow([])
            
            for key, value in section_data.items():
                if key == "title":
                    continue
                writer.writerow([key, str(value)])
            
            writer.writerow([])
            writer.writerow([])
